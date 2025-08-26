import os
import re
import pytesseract
from PIL import Image
import pdf2image
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Set Tesseract-OCR path
pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\Zemo\Desktop\Atlas Residential\Scraper\Tesseract\tesseract.exe"
)

# --- OCR Log Generator ---
def generate_ocr_log(pdf_path, output_folder, filename_base):
    try:
        images = pdf2image.convert_from_path(pdf_path, dpi=300)
        full_text = ""
        for img in images:
            text = pytesseract.image_to_string(img)
            full_text += text + "\n\n"

        os.makedirs(output_folder, exist_ok=True)
        output_path = os.path.join(output_folder, f"{filename_base}.txt")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(full_text.strip())
        return full_text
    except Exception as e:
        print(f"⚠️ Error generating OCR log for {filename_base}: {e}")
        return ""

# Base directory
base_dir = r"C:\Users\Zemo\Desktop\Atlas Residential\Scraper\Mcklenburg\Scraped File"

# Get the latest folder with prefix "Downloaded PDFs"
folders = [f for f in os.listdir(base_dir) if f.startswith("Downloaded PDFs")]
if not folders:
    raise FileNotFoundError("No 'Downloaded PDFs' folders found.")
latest_folder = max(
    folders,
    key=lambda f: datetime.strptime(f.replace("Downloaded PDFs ", ""), "%m-%d-%Y")
)
pdf_folder = os.path.join(base_dir, latest_folder)
print(f"Using folder: {pdf_folder}")

# Get the latest Excel file in that folder
excel_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".xlsx")]
if not excel_files:
    raise FileNotFoundError("No Excel files found in the folder.")
latest_excel = max(
    excel_files,
    key=lambda f: os.path.getctime(os.path.join(pdf_folder, f))
)
excel_path = os.path.join(pdf_folder, latest_excel)
print(f"Using Excel file: {excel_path}")

# Regex patterns
recording_pat = re.compile(
    r'(?:record(?:ed)?\s+on\s+)?(?:\w+\s+\d{1,2},\s+\d{4},\s+)?in\s+Book(?:\s+No\.?)?\s+(\d{3,6})[,\s]+at\s+Page[:\s]+(\d{1,6})',
    re.IGNORECASE
)


combined_pat = re.compile(
    r'Book[:\s_]*?(\d{3,6})[,\s]+Page[:\s_]*?(\d{1,6})'
    r'|Book\s+(\d{3,6})\s+Page\s+(\d{1,6})'
    r'|Book(?:\s+Mecklenburg)?[:\s]+(\d{3,6}).*?Page[:\s]+(\d{1,6})',
    re.IGNORECASE
)

# Prepare file list
pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")]
results = []
missing_book_page = []
ocr_log_dir = os.path.join(pdf_folder, "OCR  - STR Book and Page Logs")
os.makedirs(ocr_log_dir, exist_ok=True)

# Process each PDF
for pdf_file in pdf_files:
    filename_base = os.path.splitext(pdf_file)[0]
    pdf_path = os.path.join(pdf_folder, pdf_file)
    txt_path = os.path.join(ocr_log_dir, f"{filename_base}.txt")

    ocr_text = generate_ocr_log(pdf_path, ocr_log_dir, filename_base)
    if not ocr_text:
        print(f"❌ Skipping due to OCR failure: {pdf_file}")
        continue

    # Remove header to avoid incorrect Book/Page detection
    lines = ocr_text.splitlines()
    start_idx = 0
    for i, line in enumerate(lines):
        lower_line = line.strip().lower()
        if "appointment of substitute trustee" in lower_line or "substitute trustee" in lower_line:
            start_idx = i
            break
    main_text = " ".join(lines[start_idx:]).replace("\n", " ").strip()

    # Pattern matching
    book = page = None
    m_recording = recording_pat.search(main_text)
    if m_recording:
        book, page = m_recording.group(1), m_recording.group(2)
    else:
        m = combined_pat.search(main_text)
        if m:
            groups = m.groups()
            for i in range(0, len(groups), 2):
                if groups[i] and groups[i + 1]:
                    book, page = groups[i], groups[i + 1]
                    break

    if book or page:
        results.append({"file": pdf_file, "book": book, "page": page})
        print(f"{pdf_file} -> Book: {book}, Page: {page}")
    else:
        print(f"⚠️ Could not extract Book/Page from {pdf_file}")
        missing_book_page.append({
            "file": pdf_file,
            "instrument": filename_base.zfill(10),
            "date_scraped": datetime.today().strftime("%Y-%m-%d")
        })

# Update Excel
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

headers = [str(cell.value).strip().lower() for cell in ws[1]]
if 'book' not in headers:
    book_col = len(headers) + 1
    ws.cell(row=1, column=book_col, value='Book')
    page_col = book_col + 1
    ws.cell(row=1, column=page_col, value='Page')
else:
    book_col = headers.index('book') + 1
    page_col = headers.index('page') + 1

instrument_col = 1
file_row_map = {}
for row in range(2, ws.max_row + 1):
    inst = ws.cell(row=row, column=instrument_col).value
    if inst:
        key = str(inst).split('.')[0].strip().zfill(10)
        file_row_map[key] = row

for result in results:
    pdf_filename = result['file'].strip().lower()
    instrument_num = os.path.splitext(pdf_filename)[0].zfill(10)
    row = file_row_map.get(instrument_num)
    if row:
        ws.cell(row=row, column=book_col, value=result['book'])
        ws.cell(row=row, column=page_col, value=result['page'])
    else:
        print(f"No matching Instrument # for file {result['file']}")

if missing_book_page:
    missing_ws = wb.create_sheet("Missing Book Page")
    missing_ws.append(["Filename", "Instrument #", "Date Scraped"])
    for entry in missing_book_page:
        missing_ws.append([entry["file"], entry["instrument"], entry["date_scraped"]])

wb.save(excel_path)
wb.close()
print("Excel file updated successfully.")

# Upload to Google Sheets
import gspread
from oauth2client.service_account import ServiceAccountCredentials

if missing_book_page:
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key("1C6Q6iJTzO89LJRw6q2K1V-9m8NCzWegHgswfjPHanAQ").worksheet("Missing Book and Page")

    header = ["Filename", "Instrument #", "Date Scraped"]
    rows = [[entry["file"], entry["instrument"], entry["date_scraped"]] for entry in missing_book_page]
    sheet.insert_rows([header] + rows, row=1)
    print("Missing Book/Page data uploaded to Google Sheets.")
